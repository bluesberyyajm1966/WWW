"""
convert_un_extra_xlsx.py
Reads all 5 new UN Excel files and merges their data into
the existing data/un/un_data.json produced by convert_un_xlsx.py.

Place in your WWWD folder and run:
    python3 convert_un_extra_xlsx.py

Requirements:
    pip3 install openpyxl
"""

import json
import os
import sys

try:
    import openpyxl
except ImportError:
    print("Missing library. Run:  pip3 install openpyxl")
    sys.exit(1)

UN_PATH = os.path.join("data", "un", "un_data.json")

# Each entry: (filename_keyword, sheet_name, column_map)
# column_map: list of (header_keyword, field_name)
# The script does case-insensitive substring matching on headers.

FILES = [
    # ── Population by Age ─────────────────────────────────────────────────────
    {
        "keyword": "Population Data",
        "sheet":   "Estimates",
        "columns": [
            ("ISO3",                          "_ISO3_"),
            ("Year",                          "_YEAR_"),
            ("0-14",                          "popAge0to14"),
            ("15-64",                         "popAge15to64"),
            ("65",                            "popAge65plus"),
            ("Dependency",                    "dependencyRatio"),
        ],
    },

    # ── Deaths by Age (all sexes) ─────────────────────────────────────────────
    # We'll extract neonatal (under 1 month ≈ age 0) deaths as a rate if available
    {
        "keyword": "Deaths by Age",
        "sheet":   "Estimates",
        "columns": [
            ("ISO3",                          "_ISO3_"),
            ("Year",                          "_YEAR_"),
            ("Neonatal",                      "neonateMortality"),
            ("Under age 1",                   "neonateMortality"),   # fallback label
        ],
    },

    # ── Female Deaths by Age ──────────────────────────────────────────────────
    {
        "keyword": "Female Deaths Age",
        "sheet":   "Estimates",
        "columns": [
            ("ISO3",                          "_ISO3_"),
            ("Year",                          "_YEAR_"),
            ("Neonatal",                      "neonateMortalityF"),
        ],
    },

    # ── Male Deaths by Age ────────────────────────────────────────────────────
    {
        "keyword": "Male Deaths Age",
        "sheet":   "Estimates",
        "columns": [
            ("ISO3",                          "_ISO3_"),
            ("Year",                          "_YEAR_"),
            ("Neonatal",                      "neonateMortalityM"),
        ],
    },

    # ── Births by Age of Mother ───────────────────────────────────────────────
    {
        "keyword": "Births by Age",
        "sheet":   "Estimates",
        "columns": [
            ("ISO3",                          "_ISO3_"),
            ("Year",                          "_YEAR_"),
            ("Total",                         "totalBirths"),
            ("15-19",                         "birthsAge15to19"),    # teen births
            ("Adolescent",                    "adolBirthRate"),
        ],
    },
]


def find_file(keyword):
    """Find an xlsx file in the current directory whose name contains keyword."""
    for f in os.listdir("."):
        if f.endswith(".xlsx") and keyword.lower() in f.lower():
            return f
    return None


def find_header_row(all_rows):
    """Scan first 25 rows for the one containing ISO3."""
    for i, row in enumerate(all_rows[:25]):
        row_str = " ".join(str(c) for c in row if c)
        if "ISO3" in row_str or "iso3" in row_str.lower():
            return i
    return None


def map_columns(header_row, column_map):
    """Return (iso3_col, year_col, {field_name: col_idx})."""
    iso3_col   = None
    year_col   = None
    field_cols = {}

    for col_i, cell in enumerate(header_row):
        if cell is None:
            continue
        cell_str = str(cell).strip()

        for keyword, field_name in column_map:
            if keyword.lower() not in cell_str.lower():
                continue

            if field_name == "_ISO3_":
                if iso3_col is None:
                    iso3_col = col_i
            elif field_name == "_YEAR_":
                if year_col is None:
                    year_col = col_i
            else:
                if field_name not in field_cols:
                    field_cols[field_name] = col_i

    return iso3_col, year_col, field_cols


def read_file(filepath, sheet_name, column_map):
    """Read one xlsx file and return {iso3: {field: value, field_year: year}}."""
    print(f"\n  Reading: {filepath}")
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

    # Find sheet
    sheet = None
    for name in wb.sheetnames:
        if sheet_name.lower() in name.lower() or "estimate" in name.lower():
            sheet = wb[name]
            print(f"  Sheet: '{name}'")
            break
    if sheet is None:
        sheet = wb.active
        print(f"  Sheet (fallback): '{sheet.title}'")

    all_rows = list(sheet.iter_rows(values_only=True))
    print(f"  Rows: {len(all_rows)}")

    header_idx = find_header_row(all_rows)
    if header_idx is None:
        print("  ✗ Header row not found — skipping this file")
        return {}

    header = all_rows[header_idx]
    iso3_col, year_col, field_cols = map_columns(header, column_map)

    print(f"  ISO3 col: {iso3_col}  Year col: {year_col}")
    print(f"  Fields: {list(field_cols.keys())}")

    if iso3_col is None:
        print("  ✗ ISO3 column not found — skipping")
        # Print headers to help diagnose
        print("  All headers:", [str(h)[:30] for h in header if h][:15])
        return {}

    result    = {}
    best_year = {}

    for row in all_rows[header_idx + 1:]:
        if iso3_col >= len(row):
            continue
        iso3 = row[iso3_col]
        if not iso3 or len(str(iso3).strip()) != 3:
            continue
        iso3 = str(iso3).strip().upper()

        yr_int = 0
        year   = "2024"
        if year_col is not None and year_col < len(row):
            yr_raw = row[year_col]
            if yr_raw is not None:
                try:
                    yr_int = int(float(str(yr_raw)))
                    year   = str(yr_int)
                except (ValueError, TypeError):
                    pass

        if yr_int > 2025:
            continue
        if iso3 in best_year and yr_int < best_year[iso3]:
            continue
        best_year[iso3] = yr_int

        if iso3 not in result:
            result[iso3] = {}

        for field_name, col_idx in field_cols.items():
            if col_idx < len(row) and row[col_idx] is not None:
                try:
                    result[iso3][field_name]            = float(row[col_idx])
                    result[iso3][field_name + "_year"]  = year
                except (ValueError, TypeError):
                    pass

    print(f"  Countries: {len(result)}")
    return result


def main():
    print("=" * 60)
    print("  UN Extra Excel Files -> un_data.json Merger")
    print("=" * 60)

    # Load existing un_data.json
    if not os.path.exists(UN_PATH):
        print(f"\n✗ {UN_PATH} not found — run convert_un_xlsx.py first")
        sys.exit(1)
    with open(UN_PATH) as f:
        un_data = json.load(f)
    print(f"\nExisting un_data: {len(un_data)} countries")

    # Process each file
    total_merged = 0
    for spec in FILES:
        filepath = find_file(spec["keyword"])
        if not filepath:
            print(f"\n  ⚠ File not found for keyword '{spec['keyword']}' — skipping")
            continue

        new_data = read_file(filepath, spec["sheet"], spec["columns"])

        # Merge into un_data (only fill missing fields, don't overwrite)
        for iso3, fields in new_data.items():
            if iso3 not in un_data:
                un_data[iso3] = {}
            for field, value in fields.items():
                if un_data[iso3].get(field) is None:
                    un_data[iso3][field] = value
                    total_merged += 1

    print(f"\n{'='*60}")
    print(f"Total new field values merged: {total_merged}")
    print(f"Total countries in un_data:    {len(un_data)}")

    # Show USA sample
    if "USA" in un_data:
        print(f"\nSample (USA) — all fields:")
        for k, v in un_data["USA"].items():
            if not k.endswith("_year"):
                print(f"  {k:30s} = {v}")

    # Save
    os.makedirs(os.path.dirname(UN_PATH), exist_ok=True)
    with open(UN_PATH, "w") as f:
        json.dump(un_data, f, indent=2)

    print(f"\n✓ Saved to {UN_PATH}")


if __name__ == "__main__":
    main()
